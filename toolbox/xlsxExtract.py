#!/usr/bin/env python3
"""
excel_to_markdown_txt.py

Extract text from Excel workbooks (.xlsx, .xlsm) and save it as a Markdown-formatted
plain-text file. Each worksheet is rendered as a Markdown table.

Features:
- Outputs one Markdown table per worksheet with a "## Sheet: <name>" heading.
- Treats the first row as a header by default (configurable).
- Drops leading/trailing entirely empty rows and trailing empty columns (configurable).
- Preserves hyperlinks as [text](url) when possible.
- Converts line breaks in cells to <br> for valid Markdown tables.
- Attempts to format dates/times in ISO format.
- Reads formulas as cached values (data_only=True); if not calculated, the cell may be blank.

Usage:
  python excel_to_markdown_txt.py input.xlsx [-o output.txt]
Options:
  --sheet SHEET        Limit to one or more sheets by name or 1-based index (repeatable).
  --no-header          Do not treat first row as header; generate "Column N" headers instead.
  --keep-empty         Keep leading/trailing empty rows and trailing empty columns.
  --max-rows N         Limit rows per sheet (0 = no limit).
  --max-cols N         Limit columns per sheet (0 = no limit).
  --encoding ENC       Output file encoding (default: utf-8).

Notes:
- Only .xlsx/.xlsm are supported (via openpyxl).
- For large files, read_only mode is used for memory efficiency.
- Merged cells are flattened; only the top-left cell value appears.
"""

import argparse
import os
import sys
import re
from datetime import date, datetime, time

try:
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel as excel_from_serial
except ImportError as e:
    print("Missing dependency: openpyxl. Install it with:\n  pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def escape_md_table_cell(text: str) -> str:
    """
    Escape/normalize text so it is safe inside a Markdown pipe table cell.
    - Replace tabs with 4 spaces
    - Normalize CR/LF to LF and replace LF with <br>
    - Normalize non-breaking spaces to regular spaces
    """
    if text is None:
        return ""
    s = str(text)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\t", "    ")
    s = s.replace("\u00A0", " ")
    # Replace line breaks with <br> so the table stays valid
    s = s.replace("\n", "<br>")
    # Escape vertical bars so they don't split the table
    s = s.replace("|", r"\|")
    return s


def value_to_text(cell, wb_epoch) -> str:
    """
    Convert a cell value to a string suitable for Markdown.
    - Try to format dates/times in ISO format.
    - Preserve hyperlinks as [text](url) when present.
    """
    v = cell.value
    # Handle dates/times (value may already be datetime/date/time)
    try:
        is_date = bool(getattr(cell, "is_date", False))
    except Exception:
        is_date = False

    if isinstance(v, datetime):
        s = v.isoformat(sep=" ", timespec="seconds")
    elif isinstance(v, date):
        s = v.isoformat()
    elif isinstance(v, time):
        # time.isoformat can include microseconds; trim to seconds
        s = v.replace(microsecond=0).isoformat()
    elif is_date and isinstance(v, (int, float)):
        # Cell likely stores an Excel serial; convert using workbook epoch
        try:
            dt = excel_from_serial(v, wb_epoch)
            if isinstance(dt, datetime):
                s = dt.isoformat(sep=" ", timespec="seconds")
            elif isinstance(dt, date):
                s = dt.isoformat()
            elif isinstance(dt, time):
                s = dt.replace(microsecond=0).isoformat()
            else:
                s = str(v)
        except Exception:
            s = str(v)
    elif v is None:
        s = ""
    else:
        s = str(v)

    # Preserve hyperlinks if available
    url = None
    try:
        hl = getattr(cell, "hyperlink", None)
        if hl:
            url = getattr(hl, "target", None) or str(hl)
    except Exception:
        url = None

    if url:
        # Escape only the text part for pipes/newlines; URL left as-is
        s_escaped = escape_md_table_cell(s)
        return f"[{s_escaped}]({url})"

    return s


def rows_to_markdown(rows, first_row_is_header=True):
    """
    Convert a 2D list of strings to a Markdown table string.
    Assumes rows already cropped and normalized.
    """
    if not rows:
        return ""

    num_cols = max((len(r) for r in rows), default=0)
    if num_cols == 0:
        return ""

    # Normalize row lengths
    norm = []
    for r in rows:
        r2 = list(r) + [""] * (num_cols - len(r))
        norm.append(r2)

    if first_row_is_header:
        header = norm[0]
        # If header is entirely empty, synthesize generic headers
        if not any(c.strip() for c in header):
            header = [f"Column {i+1}" for i in range(num_cols)]
            body = norm
        else:
            body = norm[1:]
    else:
        header = [f"Column {i+1}" for i in range(num_cols)]
        body = norm

    sep = ["---"] * num_cols

    lines = []
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(sep) + " |")
    for r in body:
        lines.append("| " + " | ".join(r) + " |")

    return "\n".join(lines)


def crop_rows_and_cols(rows, keep_empty, max_rows=0, max_cols=0):
    """
    - Drop leading/trailing completely empty rows and trailing empty columns (unless keep_empty).
    - Apply optional max_rows/max_cols limits.
    """
    if not rows:
        return []

    # Determine content bounds
    def row_is_empty(r):
        return not any((c.strip() if isinstance(c, str) else str(c).strip()) for c in r)

    last_non_empty_col = -1
    last_non_empty_row = -1

    for ri, r in enumerate(rows):
        for ci, c in enumerate(r):
            if (c.strip() if isinstance(c, str) else str(c).strip()):
                if ci > last_non_empty_col:
                    last_non_empty_col = ci
                last_non_empty_row = max(last_non_empty_row, ri)

    # If all empty
    if last_non_empty_row == -1:
        return []

    # Drop leading empty rows
    start_row = 0
    if not keep_empty:
        while start_row <= last_non_empty_row and row_is_empty(rows[start_row]):
            start_row += 1

    # Drop trailing empty rows
    end_row = last_non_empty_row if not keep_empty else len(rows) - 1

    # Determine last useful column
    end_col = (len(rows[0]) - 1)
    if not keep_empty and last_non_empty_col >= 0:
        end_col = last_non_empty_col

    # Apply max limits
    if max_rows and max_rows > 0:
        end_row = min(end_row, start_row + max_rows - 1)
    if max_cols and max_cols > 0:
        end_col = min(end_col, max_cols - 1)

    cropped = []
    for ri in range(start_row, end_row + 1):
        row = rows[ri]
        # Pad row to at least end_col+1
        if len(row) <= end_col:
            row = row + [""] * (end_col + 1 - len(row))
        cropped.append(row[: end_col + 1])

    return cropped


def worksheet_to_rows(ws, wb_epoch, max_rows=0, max_cols=0):
    """
    Read a worksheet and return a list of rows, where each row is a list of
    Markdown-safe strings (pipes/newlines escaped/replaced).
    Applies optional per-sheet row/column limits during reading to save memory.
    """
    rows = []
    count_rows = 0
    for row in ws.iter_rows(values_only=False):
        if max_rows and count_rows >= max_rows:
            break
        out_row = []
        count_cols = 0
        for cell in row:
            if max_cols and count_cols >= max_cols:
                break
            s = value_to_text(cell, wb_epoch)
            s = escape_md_table_cell(s)
            out_row.append(s)
            count_cols += 1
        rows.append(out_row)
        count_rows += 1
    return rows


def parse_sheet_selectors(selectors, wb):
    """
    selectors: list of names or 1-based indices (as strings)
    Returns a list of actual sheet names present in the workbook.
    """
    if not selectors:
        return list(wb.sheetnames)

    result = []
    for sel in selectors:
        sel = str(sel).strip()
        if not sel:
            continue
        # Try numeric index (1-based)
        if sel.isdigit():
            idx = int(sel)
            if 1 <= idx <= len(wb.sheetnames):
                result.append(wb.sheetnames[idx - 1])
            else:
                print(f"Warning: sheet index out of range: {sel}", file=sys.stderr)
        else:
            # Match by name (case-sensitive)
            if sel in wb.sheetnames:
                result.append(sel)
            else:
                print(f"Warning: sheet not found: {sel}", file=sys.stderr)
    # Deduplicate while preserving order
    seen = set()
    uniq = []
    for name in result:
        if name not in seen:
            uniq.append(name)
            seen.add(name)
    return uniq


def main(inputPath):
    ap = argparse.ArgumentParser(description="Extract text from Excel (.xlsx/.xlsm) into Markdown-formatted .txt")
    ap.add_argument("--sheet", action="append", dest="sheets",
                    help="Sheet name or 1-based index to export (repeatable). Default: all sheets.")
    ap.add_argument("--no-header", action="store_true", help="Do not treat first row as header; synthesize 'Column N' headers.")
    ap.add_argument("--keep-empty", action="store_true", help="Keep leading/trailing empty rows and trailing empty columns.")
    ap.add_argument("--max-rows", type=int, default=0, help="Limit rows per sheet (0 = no limit).")
    ap.add_argument("--max-cols", type=int, default=0, help="Limit columns per sheet (0 = no limit).")
    ap.add_argument("--encoding", default="utf-8", help="Output file encoding (default: utf-8).")
    args = ap.parse_args()

    finalContent = ""

    input_path = inputPath
    if not os.path.isfile(input_path):
        print(f"Error: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    _, ext = os.path.splitext(input_path)
    if ext.lower() not in (".xlsx", ".xlsm"):
        print("Error: Only .xlsx and .xlsm files are supported.", file=sys.stderr)
        sys.exit(1)

    try:
        # Use read_only for large files; data_only to get cached formula results.
        wb = load_workbook(input_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"Failed to open workbook: {e}", file=sys.stderr)
        sys.exit(1)

    try:
        selected_sheets = parse_sheet_selectors(args.sheets, wb)
        if not selected_sheets:
            print("No matching sheets found.", file=sys.stderr)
            sys.exit(1)

        out_lines = []
        # Add a top-level title with the file name
        out_lines.append("")

        wb_epoch = getattr(wb, "epoch", datetime(1899, 12, 30))

        for sheet_name in selected_sheets:
            ws = wb[sheet_name]

            rows = worksheet_to_rows(ws, wb_epoch, max_rows=args.max_rows or 0, max_cols=args.max_cols or 0)
            rows = crop_rows_and_cols(rows, keep_empty=args.keep_empty, max_rows=0, max_cols=0)

            # Skip entirely empty sheets
            if not rows:
                # Still include a heading and note for visibility
                out_lines.append(f"## Sheet: {sheet_name}")
                out_lines.append("")
                out_lines.append("_(No data)_")
                out_lines.append("")
                continue

            out_lines.append(f"## Sheet: {sheet_name}")
            out_lines.append("")
            table_md = rows_to_markdown(rows, first_row_is_header=(not args.no_header))
            out_lines.append(table_md)
            out_lines.append("")

        content = "\n".join(out_lines).rstrip() + "\n"

        finalContent = content

    except Exception as e:
        print(f"Conversion failed: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        try:
            wb.close()
        except Exception:
            pass
    
    return finalContent

